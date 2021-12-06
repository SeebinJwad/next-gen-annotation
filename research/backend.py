import pandas as pd
import openpyxl
import os

# TODO: flip ud_a and ud_b in compare (more intuitive ud_a - ud_b)
chromosome_names = ['1', '2', '3', '4', '5', '6', '7', '8', '9', '10', '11', '12', '13', '14', '15', '16', '17', '18',
                    '19', '20', '21', "X", "Y"]


def generate_file_names(sample_id):
    file_names = []
    for x in range(23):
        file_names.append("anno1.fixed.chr" + chromosome_names[x] + ".genes." + sample_id + ".hg38_multianno.txt")
    return file_names


def pass_filter(file_df, dp_threshold=0, af_threshold=0.0, ref_gene=False):
    rows = []
    for x in range(len(file_df['Otherinfo11'].copy().index)):
        dp = False
        af = False
        for i in file_df.loc[x, 'Otherinfo11'].split(";"):
            if i[:3] == "DP=" and int(i[3:]) > dp_threshold:
                dp = True
            if i[:3] == "AF=" and float(i[3:]) > af_threshold:
                af = True
        if (dp and af and file_df.loc[x, 'Otherinfo10'] == "PASS" and not ref_gene) or \
                (dp and af and file_df.loc[x, 'Otherinfo10'] == "PASS" and ref_gene and (file_df.loc[x, 'Func.refGene'] == "exonic" or file_df.loc[x, 'Func.refGene'] == "splicing")):
            rows.append(x)

    return rows


def combine(df, first_pass):
    start = df["Start"].tolist()
    start = [str(x) for x in start]
    end = df["End"].tolist()
    end = [str(x) for x in end]
    ref = df["Ref"].tolist()
    alt = df["Alt"].tolist()
    generated_ud = [j + k + l + m for j, k, l, m in zip(start, end, ref, alt)]
    filtered_ud = []
    for passed_index in first_pass:
        filtered_ud.append(generated_ud[passed_index])
    return dict(zip(filtered_ud, first_pass))


def compare(ud_a, ud_b):
    # the uds from ud_b that are not in ud_a
    disjoints = []
    for k in range(len(ud_b)):
        if list(ud_b.keys())[k] not in list(ud_a.keys()):
            disjoints.append(list(ud_b.values())[k] + 2)
    return disjoints


def get_row(file, chromosome_index):
    output = []
    for x in range(len(file)):
        cell2 = file.iloc[x, chromosome_index]
        # > 0 checks whether the cell is NaN or not
        if cell2 > 0:
            output.append(int(cell2))
    return output


def print_xlsx(sample_id, chromosome_index, rows):
    file_name = generate_file_names(sample_id)[chromosome_index]
    df = pd.read_csv(r"C:\\Users\\kazij\\Documents\\Cisplatin_Resistance_Variant_Analysis\\annovar-full\\" + file_name, delimiter="\t")
    output = []
    for x in rows:
        output.append([sample_id, df.loc[x-2, 'Chr'], df.loc[x-2, 'Start'], df.loc[x-2, 'End'], df.loc[x-2, 'Ref'], df.loc[x-2, 'Alt'],
                       df.loc[x-2, 'Func.refGene'], df.loc[x-2, 'Gene.refGene'], df.loc[x-2, 'GeneDetail.refGene'], df.loc[x-2, 'ExonicFunc.refGene'],
                       df.loc[x-2, 'AAChange.refGene'], df.loc[x-2, 'Otherinfo1'], df.loc[x-2, 'Otherinfo2'], df.loc[x-2, 'Otherinfo3'],
                       df.loc[x - 2, 'Otherinfo4'], df.loc[x-2, 'Otherinfo5'], df.loc[x-2, 'Otherinfo6'], df.loc[x-2, 'Otherinfo7'],
                       df.loc[x - 2, 'Otherinfo8'], df.loc[x-2, 'Otherinfo9'], df.loc[x-2, 'Otherinfo10'], df.loc[x-2, 'Otherinfo11'],
                       df.loc[x-2, 'Otherinfo12'], df.loc[x-2, 'Otherinfo13']])
    return output


# minuend one being subtracted, subtrahend subtracting
def main_comp(minuend_id, subtrahend_id, outputs="ALL", DP_minuend=0, DP_subtrahend=0, AF_minuend=0, AF_subtrahend=0, QC_minuend=True, QC_subtrahend=True):
    # OPPOSITE FILE NAMES
    variant_sample_id = minuend_id # "VE_22241"
    parent_sample_id = subtrahend_id # "VE_22251"
    variant_files = generate_file_names(parent_sample_id)
    parent_files = generate_file_names(variant_sample_id)
    wb = openpyxl.Workbook()

    ws1 = wb.create_sheet("parent")
    ws1.title = "parent"

    ws2 = wb.create_sheet("variant")
    ws2.title = "variant"

    for x in range(23):
        var_df_inp = pd.read_csv(r"C:\\Users\\kazij\\Documents\\Cisplatin_Resistance_Variant_Analysis\\annovar-full\\" + variant_files[x], delimiter="\t")
        par_df_inp = pd.read_csv(r"C:\\Users\\kazij\\Documents\\Cisplatin_Resistance_Variant_Analysis\\annovar-full\\" + parent_files[x], delimiter="\t")
        # subtracting 30 .15, subtracted from 25 .1
        # dp af thresholds for pass_filter are 0 if dp af should be ignored
        variant_ud = combine(var_df_inp, pass_filter(var_df_inp, DP_minuend, AF_minuend))
        parent_ud = combine(par_df_inp, pass_filter(par_df_inp, DP_subtrahend, AF_subtrahend))
        # print(compare(parent_ud, variant_ud))
        output = compare(parent_ud, variant_ud)
        sheet = wb["parent"]
        for g in range(len(output)):
            sheet.cell(row=g+2, column=x+1).value = output[g]
        variant_ud = combine(var_df_inp, pass_filter(var_df_inp, DP_subtrahend, AF_subtrahend))
        # 35, .15, True
        parent_ud = combine(par_df_inp, pass_filter(par_df_inp, DP_minuend, AF_minuend))
        output = compare(variant_ud, parent_ud)
        sheet = wb["variant"]
        for g in range(len(output)):
            sheet.cell(row=g+2, column=x+1).value = output[g]
        # add chromosome name headers to both sheets
        wb["parent"].cell(row=1, column=x+1).value = chromosome_names[x]
        wb["variant"].cell(row=1, column=x + 1).value = chromosome_names[x]
    # move save location to hn31 results
    wb.save("result4.xlsx")

    # GET ROW PART
    variant_sample_id = variant_sample_id
    parent_sample_id = parent_sample_id
    variant = pd.read_excel("result4.xlsx", sheet_name="variant")
    parent = pd.read_excel("result4.xlsx", sheet_name="parent")
    wb2 = openpyxl.Workbook()

    sheet2p = wb2.create_sheet("parent_rows")
    sheet2p.title = "parent_rows"

    sheet2v = wb2.create_sheet("variant_rows")
    sheet2v.title = "variant_rows"

    # OUTPUT TO VARIANT
    row_count = 0
    for x in range(23):
        output_list = print_xlsx(variant_sample_id, x, get_row(variant, x))
        for i in output_list:
            row_count += 1
            for j in range(24):
                sheet2v.cell(row=row_count + 1, column=j + 1).value = i[j]

    # OUTPUT TO PARENT
    row_count = 0
    for x in range(23):
        output_list = print_xlsx(parent_sample_id, x, get_row(parent, x))
        for i in output_list:
            row_count += 1
            for j in range(24):
                sheet2p.cell(row=row_count + 1, column=j+1).value = i[j]
    headers = ['Sample ID', 'Chr', 'Start', 'End', 'Ref', 'Alt', 'Func.refGene', 'Gene.refGene', 'GeneDetail.refGene', 'ExonicFunc.refGene',
               'AAChange.refGene', 'Otherinfo1', 'Otherinfo2', 'Otherinfo3', 'Otherinfo4', 'Otherinfo5', 'Otherinfo6', 'Otherinfo7',
               'Otherinfo8', 'Otherinfo9', 'Otherinfo10', 'Otherinfo11', 'Otherinfo12', 'Otherinfo13']
    for j in range(24):
        sheet2p.cell(row=1, column=j+1).value = headers[j]
        sheet2v.cell(row=1, column=j+1).value = headers[j]
    wb2.save(r"C:\\Users\\kazij\\Documents\\Cisplatin_Resistance_Variant_Analysis\\results\\" + parent_sample_id + "_" + variant_sample_id + ".xlsx")
    os.remove("result4.xlsx")


"""main_comp("VE_22252", "VE_22251")
main_comp("VE_22253", "VE_22251")
main_comp("VE_22254", "VE_22251")
main_comp("VE_22255", "VE_22251")

main_comp("VE_22247", "VE_22246")
main_comp("VE_22248", "VE_22246")
main_comp("VE_22249", "VE_22246")
main_comp("VE_22250", "VE_22246")"""
